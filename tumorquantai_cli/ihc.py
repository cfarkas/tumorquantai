"""Deterministic breast-IHC quantification and agreement reporting.

This module is deliberately separate from the HistoPLUS H&E cell-typing
worker.  It provides a package-native, marker-aware analysis for brightfield
Hematoxylin/DAB patches:

* ER, PR, and Ki-67: watershed-segmented nuclei and DAB-positive percentages.
* HER2: a reproducible membrane-proxy measurement and 0/1+/2+/3+ pre-score.
* Cohort aggregation, visual QC, resumable per-patch records, and comparison
  with pathologist values through an explicit private linkage table.

The measurements are research proxies.  They are not clinical assay results,
do not identify invasive-tumour cells, and must not be used for patient care.
"""

from __future__ import annotations

import csv
import fnmatch
import gzip
import hashlib
import html
import json
import math
import os
import tempfile
import traceback
import zipfile
from collections import Counter, defaultdict
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Any, Iterable, Mapping, Sequence

try:
    import numpy as np
except ModuleNotFoundError:  # Install must work before dependencies exist.
    np = None  # type: ignore[assignment]


IHC_SCHEMA_VERSION = "tumorquantai_ihc_v1"
IHC_ENGINE_VERSION = "hdab-watershed-membrane-proxy-v1"
PUBLIC_DATASET_RECORD = "21797920"
PUBLIC_DATASET_DOI = "10.5281/zenodo.21797920"
IHC_MARKERS = ("ER", "PR", "HER2", "Ki-67")
NUCLEAR_MARKERS = frozenset({"ER", "PR", "Ki-67"})
MANIFEST_REQUIRED_COLUMNS = frozenset(
    {
        "case_alias",
        "patch_alias",
        "marker",
        "public_path",
        "microns_per_pixel",
        "width",
        "height",
        "decoded_rgb_sha256",
    }
)

# Ruifrok-Johnston HED separation matrix.  Multiplication is performed on
# natural-log optical density, so the resulting H and DAB concentrations use
# optical-density-like units rather than scikit-image's normalized units.
HED_FROM_RGB = (
    np.asarray(
        [
            [1.87798274, -1.00767869, -0.55611582],
            [-0.06590806, 1.13473037, -0.13552180],
            [-0.60190736, -0.48041419, 1.57358807],
        ],
        dtype=np.float32,
    )
    if np is not None
    else None
)


class IHCError(RuntimeError):
    """Expected, user-facing IHC analysis error."""


def utc_now() -> str:
    return (
        datetime.now(timezone.utc)
        .replace(microsecond=0)
        .isoformat()
        .replace("+00:00", "Z")
    )


def _atomic_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="") as handle:
            handle.write(value)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        try:
            os.unlink(temporary)
        except FileNotFoundError:
            pass


def _atomic_json(path: Path, value: Mapping[str, Any]) -> None:
    _atomic_text(path, json.dumps(value, indent=2, sort_keys=True) + "\n")


def _write_csv(
    path: Path, rows: Sequence[Mapping[str, Any]], fields: Sequence[str]
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle, fieldnames=list(fields), extrasaction="ignore"
            )
            writer.writeheader()
            writer.writerows(rows)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        try:
            os.unlink(temporary)
        except FileNotFoundError:
            pass


def require_image_dependencies() -> tuple[Any, Any, Any, Any, Any, Any, Any]:
    """Import the optional scientific stack only for the IHC command."""
    missing: list[str] = []
    try:
        import cv2
    except ModuleNotFoundError:
        cv2 = None
        missing.append("opencv-python-headless")
    try:
        import tifffile
    except ModuleNotFoundError:
        tifffile = None
        missing.append("tifffile")
    try:
        from PIL import Image, ImageDraw
    except ModuleNotFoundError:
        Image = ImageDraw = None
        missing.append("Pillow")
    try:
        from scipy import ndimage
    except ModuleNotFoundError:
        ndimage = None
        missing.append("scipy")
    try:
        from skimage import feature, measure, morphology, segmentation
    except ModuleNotFoundError:
        feature = measure = morphology = segmentation = None
        missing.append("scikit-image")
    if missing:
        raise IHCError(
            "IHC quantification dependencies are missing: "
            + ", ".join(sorted(set(missing)))
            + ". Rerun 'tumorquantai install --docker' from this checkout."
        )
    return (
        cv2,
        tifffile,
        Image,
        ImageDraw,
        ndimage,
        feature,
        (
            measure,
            morphology,
            segmentation,
        ),
    )


@dataclass(frozen=True)
class IHCConfig:
    """Versioned, explicit settings for one deterministic analysis."""

    weak_dab_od: float = 0.20
    moderate_dab_od: float = 0.40
    strong_dab_od: float = 0.60
    minimum_nucleus_area_um2: float = 12.0
    maximum_nucleus_area_um2: float = 420.0
    maximum_nucleus_eccentricity: float = 0.985
    minimum_peak_distance_um: float = 2.4
    minimum_nuclear_signal_od: float = 0.12
    maximum_nuclear_signal_od: float = 0.48
    cell_expansion_um: float = 7.0
    complete_membrane_fraction: float = 0.75
    minimum_tissue_fraction: float = 0.01
    minimum_cells_for_score: int = 100
    er_pr_positive_percent: float = 1.0
    ki67_high_percent: float = 20.0
    her2_positive_cell_percent: float = 10.0

    def validate(self) -> None:
        values = asdict(self)
        for key, value in values.items():
            if isinstance(value, bool):
                continue
            if not isinstance(value, (int, float)) or not math.isfinite(float(value)):
                raise IHCError(f"Invalid non-finite IHC setting: {key}")
        if not 0 < self.weak_dab_od < self.moderate_dab_od < self.strong_dab_od:
            raise IHCError("DAB thresholds must be strictly increasing and positive")
        if not 0 < self.minimum_nucleus_area_um2 < self.maximum_nucleus_area_um2:
            raise IHCError("Nucleus-area bounds are invalid")
        if not 0 < self.maximum_nucleus_eccentricity <= 1:
            raise IHCError("Maximum nucleus eccentricity must be in (0, 1]")
        if not 0 < self.complete_membrane_fraction <= 1:
            raise IHCError("Complete-membrane fraction must be in (0, 1]")
        if self.minimum_cells_for_score < 1:
            raise IHCError("Minimum cells for score must be at least 1")

    def signature(self) -> str:
        payload = {
            "engine": IHC_ENGINE_VERSION,
            "schema": IHC_SCHEMA_VERSION,
            "settings": asdict(self),
        }
        return hashlib.sha256(
            json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
        ).hexdigest()


@dataclass(frozen=True)
class PatchRecord:
    case_alias: str
    patch_alias: str
    marker: str
    public_path: str
    microns_per_pixel: float
    width: int
    height: int
    expected_decoded_rgb_sha256: str
    source_file: str | None
    archive_file: str | None
    archive_member: str | None

    @property
    def output_relative(self) -> Path:
        return Path("patches") / self.case_alias / self.patch_alias


def _safe_public_path(value: str, row_number: int) -> str:
    candidate = PurePosixPath(value)
    if not value or candidate.is_absolute() or ".." in candidate.parts:
        raise IHCError(f"Unsafe public_path at manifest row {row_number}")
    if candidate.suffix.casefold() not in {".tif", ".tiff"}:
        raise IHCError(f"Non-TIFF public_path at manifest row {row_number}")
    return candidate.as_posix()


def _positive_float(value: Any, label: str, row_number: int) -> float:
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise IHCError(f"Invalid {label} at manifest row {row_number}") from exc
    if not math.isfinite(result) or result <= 0:
        raise IHCError(f"Invalid {label} at manifest row {row_number}")
    return result


def _positive_int(value: Any, label: str, row_number: int) -> int:
    try:
        result = int(value)
    except (TypeError, ValueError) as exc:
        raise IHCError(f"Invalid {label} at manifest row {row_number}") from exc
    if result <= 0:
        raise IHCError(f"Invalid {label} at manifest row {row_number}")
    return result


def load_patch_manifest(
    manifest: Path,
    input_root: Path,
    *,
    markers: Iterable[str] = IHC_MARKERS,
    include: str = "*",
    exclude: str = "",
) -> tuple[list[PatchRecord], list[dict[str, str]]]:
    """Resolve manifest rows to extracted TIFFs or case ZIP members."""
    manifest = manifest.expanduser().resolve()
    input_root = input_root.expanduser().resolve()
    if not manifest.is_file():
        raise IHCError(f"Patch manifest does not exist: {manifest}")
    if not input_root.is_dir():
        raise IHCError(f"IHC input directory does not exist: {input_root}")
    selected_markers = set(markers)
    unknown = selected_markers - set(IHC_MARKERS)
    if unknown:
        raise IHCError("Unknown IHC marker(s): " + ", ".join(sorted(unknown)))

    records: list[PatchRecord] = []
    unavailable: list[dict[str, str]] = []
    seen_patch_aliases: set[str] = set()
    with manifest.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fields = set(reader.fieldnames or [])
        missing = MANIFEST_REQUIRED_COLUMNS - fields
        if missing:
            raise IHCError(
                "Patch manifest is missing required columns: "
                + ", ".join(sorted(missing))
            )
        for row_number, row in enumerate(reader, start=2):
            marker = str(row.get("marker", "")).strip()
            if marker not in selected_markers:
                continue
            case_alias = str(row.get("case_alias", "")).strip()
            patch_alias = str(row.get("patch_alias", "")).strip()
            if not fnmatch.fnmatchcase(case_alias, include):
                continue
            if exclude and fnmatch.fnmatchcase(case_alias, exclude):
                continue
            if not _is_public_case_alias(case_alias):
                raise IHCError(f"Invalid case_alias at manifest row {row_number}")
            if not patch_alias.startswith("TQA_PATCH_"):
                raise IHCError(f"Invalid patch_alias at manifest row {row_number}")
            if patch_alias in seen_patch_aliases:
                raise IHCError(f"Duplicate patch_alias at manifest row {row_number}")
            seen_patch_aliases.add(patch_alias)
            public_path = _safe_public_path(str(row.get("public_path", "")), row_number)
            extracted = input_root / PurePosixPath(public_path)
            archive_candidates = (
                input_root / f"{case_alias}.zip",
                input_root / "downloads" / f"{case_alias}.zip",
            )
            source_file: str | None = None
            archive_file: str | None = None
            archive_member: str | None = None
            if extracted.is_file():
                source_file = str(extracted)
            else:
                archive = next(
                    (path for path in archive_candidates if path.is_file()), None
                )
                if archive is not None:
                    archive_file = str(archive)
                    archive_member = public_path
                else:
                    unavailable.append(
                        {
                            "case_alias": case_alias,
                            "patch_alias": patch_alias,
                            "marker": marker,
                            "public_path": public_path,
                        }
                    )
                    continue
            expected_hash = str(row.get("decoded_rgb_sha256", "")).strip().lower()
            if len(expected_hash) != 64 or any(
                c not in "0123456789abcdef" for c in expected_hash
            ):
                raise IHCError(
                    f"Invalid decoded RGB SHA-256 at manifest row {row_number}"
                )
            records.append(
                PatchRecord(
                    case_alias=case_alias,
                    patch_alias=patch_alias,
                    marker=marker,
                    public_path=public_path,
                    microns_per_pixel=_positive_float(
                        row.get("microns_per_pixel"), "microns_per_pixel", row_number
                    ),
                    width=_positive_int(row.get("width"), "width", row_number),
                    height=_positive_int(row.get("height"), "height", row_number),
                    expected_decoded_rgb_sha256=expected_hash,
                    source_file=source_file,
                    archive_file=archive_file,
                    archive_member=archive_member,
                )
            )
    records.sort(
        key=lambda row: (row.case_alias, IHC_MARKERS.index(row.marker), row.patch_alias)
    )
    return records, unavailable


def decoded_rgb_sha256(rgb: np.ndarray) -> str:
    """Return the dataset's domain-separated decoded-pixel digest."""
    if rgb.ndim != 3 or rgb.shape[-1] != 3 or not rgb.flags.c_contiguous:
        raise IHCError("Decoded pixel digest requires contiguous RGB")
    if rgb.dtype != np.uint8:
        raise IHCError("IHC quantification currently requires uint8 RGB TIFFs")
    digest = hashlib.sha256()
    digest.update(b"TumorQuantAI decoded RGB sha256 v1\x00")
    digest.update(str(rgb.shape[0]).encode("ascii"))
    digest.update(b"x")
    digest.update(str(rgb.shape[1]).encode("ascii"))
    digest.update(b"x3\x00uint8\x00")
    digest.update(memoryview(rgb).cast("B"))
    return digest.hexdigest()


def load_patch_rgb(record: PatchRecord) -> np.ndarray:
    """Decode one manifest TIFF without requiring case-archive extraction."""
    _cv2, tifffile, *_rest = require_image_dependencies()
    try:
        if record.source_file is not None:
            with tifffile.TiffFile(record.source_file) as tif:
                if len(tif.pages) != 1:
                    raise IHCError("IHC TIFF must contain exactly one page")
                value = tif.pages[0].asarray()
        else:
            if record.archive_file is None or record.archive_member is None:
                raise IHCError("Patch source locator is incomplete")
            with zipfile.ZipFile(record.archive_file, "r") as archive:
                info = archive.getinfo(record.archive_member)
                if info.is_dir():
                    raise IHCError("Manifest points to an archive directory")
                with archive.open(info, "r") as member:
                    with tifffile.TiffFile(member) as tif:
                        if len(tif.pages) != 1:
                            raise IHCError("IHC TIFF must contain exactly one page")
                        value = tif.pages[0].asarray()
    except IHCError:
        raise
    except (OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
        raise IHCError(
            f"Could not decode {record.case_alias}/{record.patch_alias}: {exc}"
        ) from exc
    rgb = np.asarray(value)
    if rgb.ndim != 3:
        raise IHCError("IHC TIFF did not decode to a three-dimensional array")
    if rgb.shape[-1] == 3:
        pass
    elif rgb.shape[0] == 3:
        rgb = np.moveaxis(rgb, 0, -1)
    else:
        raise IHCError("IHC TIFF did not decode to exactly three RGB channels")
    rgb = np.ascontiguousarray(rgb)
    if rgb.dtype != np.uint8:
        raise IHCError("IHC quantification currently requires 8-bit RGB TIFFs")
    if rgb.shape[:2] != (record.height, record.width):
        raise IHCError(
            "Decoded TIFF geometry differs from the public manifest "
            f"({rgb.shape[1]}x{rgb.shape[0]} versus {record.width}x{record.height})"
        )
    return rgb


def separate_hematoxylin_dab(rgb: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    """Separate H and DAB concentrations using natural-log optical density."""
    if rgb.dtype != np.uint8 or rgb.ndim != 3 or rgb.shape[-1] != 3:
        raise IHCError("Stain separation requires an 8-bit RGB image")
    optical = rgb.astype(np.float32)
    optical += 1.0
    optical *= 1.0 / 256.0
    np.log(optical, out=optical)
    optical *= -1.0
    hematoxylin = (
        optical[..., 0] * HED_FROM_RGB[0, 0]
        + optical[..., 1] * HED_FROM_RGB[1, 0]
        + optical[..., 2] * HED_FROM_RGB[2, 0]
    )
    dab = (
        optical[..., 0] * HED_FROM_RGB[0, 2]
        + optical[..., 1] * HED_FROM_RGB[1, 2]
        + optical[..., 2] * HED_FROM_RGB[2, 2]
    )
    np.maximum(hematoxylin, 0, out=hematoxylin)
    np.maximum(dab, 0, out=dab)
    return hematoxylin.astype(np.float32, copy=False), dab.astype(
        np.float32, copy=False
    )


def tissue_mask(rgb: np.ndarray, mpp: float) -> np.ndarray:
    cv2, *_rest = require_image_dependencies()
    minimum = rgb.min(axis=2)
    maximum = rgb.max(axis=2)
    gray = cv2.cvtColor(rgb, cv2.COLOR_RGB2GRAY)
    mask = (gray < 245) & (((maximum.astype(np.int16) - minimum) > 7) | (gray < 220))
    radius = max(1, int(round(0.8 / mpp)))
    radius = min(radius, 7)
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (2 * radius + 1,) * 2)
    cleaned = cv2.morphologyEx(mask.astype(np.uint8), cv2.MORPH_CLOSE, kernel)
    return cleaned.astype(bool)


@dataclass
class SegmentedNuclei:
    labels: np.ndarray
    label_ids: np.ndarray
    centroid_y: np.ndarray
    centroid_x: np.ndarray
    area_um2: np.ndarray
    mean_hematoxylin_od: np.ndarray
    mean_dab_od: np.ndarray
    nuclear_threshold_od: float


def segment_nuclei(
    hematoxylin: np.ndarray,
    dab: np.ndarray,
    tissue: np.ndarray,
    marker: str,
    mpp: float,
    config: IHCConfig,
) -> SegmentedNuclei:
    """Segment counterstained nuclei with deterministic marker-controlled watershed."""
    cv2, _tifffile, _Image, _ImageDraw, ndimage, feature, sk = (
        require_image_dependencies()
    )
    measure, morphology, segmentation = sk
    cv2.setNumThreads(1)
    signal = hematoxylin if marker == "HER2" else np.maximum(hematoxylin, dab)
    sigma_px = max(0.6, min(2.5, 0.55 / mpp))
    blurred = cv2.GaussianBlur(
        signal,
        ksize=(0, 0),
        sigmaX=sigma_px,
        sigmaY=sigma_px,
        borderType=cv2.BORDER_REPLICATE,
    )
    sample = blurred[::4, ::4]
    sample = sample[tissue[::4, ::4] & np.isfinite(sample) & (sample > 0.025)]
    if sample.size >= 64:
        sample_8 = np.clip(sample * (255.0 / 2.5), 0, 255).astype(np.uint8)
        threshold_8, _ = cv2.threshold(
            sample_8, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU
        )
        adaptive = float(threshold_8) * (2.5 / 255.0)
    else:
        adaptive = config.minimum_nuclear_signal_od
    threshold = float(
        np.clip(
            adaptive,
            config.minimum_nuclear_signal_od,
            config.maximum_nuclear_signal_od,
        )
    )
    mask = (blurred >= threshold) & tissue
    minimum_pixels = max(3, int(round(config.minimum_nucleus_area_um2 / (mpp * mpp))))
    mask = morphology.remove_small_objects(mask, max(3, minimum_pixels // 3))
    mask = morphology.remove_small_holes(mask, max(3, minimum_pixels // 2))
    distance = ndimage.distance_transform_edt(mask)
    minimum_distance = max(2, int(round(config.minimum_peak_distance_um / mpp)))
    coordinates = feature.peak_local_max(
        distance,
        min_distance=minimum_distance,
        threshold_abs=max(1.0, 0.7 / mpp),
        labels=mask,
        exclude_border=False,
    )
    markers = np.zeros(mask.shape, dtype=np.int32)
    if coordinates.size:
        markers[tuple(coordinates.T)] = np.arange(
            1, len(coordinates) + 1, dtype=np.int32
        )
        labels = segmentation.watershed(
            -distance, markers, mask=mask, compactness=0.001
        )
    else:
        labels = measure.label(mask, connectivity=1)
    if int(labels.max()) == 0:
        empty_float = np.empty(0, dtype=np.float64)
        return SegmentedNuclei(
            labels=np.zeros(mask.shape, dtype=np.int32),
            label_ids=np.empty(0, dtype=np.int32),
            centroid_y=empty_float,
            centroid_x=empty_float,
            area_um2=empty_float,
            mean_hematoxylin_od=empty_float,
            mean_dab_od=empty_float,
            nuclear_threshold_od=threshold,
        )

    properties = measure.regionprops_table(
        labels,
        properties=("label", "area", "centroid", "eccentricity"),
    )
    area_um2 = properties["area"].astype(np.float64) * (mpp * mpp)
    keep = (
        (area_um2 >= config.minimum_nucleus_area_um2)
        & (area_um2 <= config.maximum_nucleus_area_um2)
        & (properties["eccentricity"] <= config.maximum_nucleus_eccentricity)
    )
    retained_original = properties["label"][keep].astype(np.int64)
    mapping = np.zeros(int(labels.max()) + 1, dtype=np.int32)
    mapping[retained_original] = np.arange(
        1, len(retained_original) + 1, dtype=np.int32
    )
    clean_labels = mapping[labels]
    label_ids = np.arange(1, len(retained_original) + 1, dtype=np.int32)
    flattened = clean_labels.ravel()
    counts = np.bincount(flattened, minlength=len(label_ids) + 1).astype(np.float64)
    h_sums = np.bincount(
        flattened,
        weights=hematoxylin.ravel(),
        minlength=len(label_ids) + 1,
    )
    d_sums = np.bincount(
        flattened,
        weights=dab.ravel(),
        minlength=len(label_ids) + 1,
    )
    denominator = np.maximum(counts[1:], 1.0)
    return SegmentedNuclei(
        labels=clean_labels,
        label_ids=label_ids,
        centroid_y=properties["centroid-0"][keep].astype(np.float64),
        centroid_x=properties["centroid-1"][keep].astype(np.float64),
        area_um2=area_um2[keep],
        mean_hematoxylin_od=h_sums[1:] / denominator,
        mean_dab_od=d_sums[1:] / denominator,
        nuclear_threshold_od=threshold,
    )


def _intensity_classes(values: np.ndarray, config: IHCConfig) -> np.ndarray:
    classes = np.zeros(values.shape, dtype=np.uint8)
    classes[values >= config.weak_dab_od] = 1
    classes[values >= config.moderate_dab_od] = 2
    classes[values >= config.strong_dab_od] = 3
    return classes


def her2_membrane_measurements(
    nuclei: SegmentedNuclei,
    dab: np.ndarray,
    mpp: float,
    config: IHCConfig,
) -> dict[str, np.ndarray]:
    """Measure DAB on an expanded-nucleus boundary as a membrane proxy."""
    *_head, sk = require_image_dependencies()
    _measure, _morphology, segmentation = sk
    count = len(nuclei.label_ids)
    if count == 0:
        empty = np.empty(0, dtype=np.float64)
        return {
            "mean_membrane_dab_od": empty,
            "positive_membrane_fraction": empty,
            "complete_membrane_proxy": np.empty(0, dtype=bool),
            "intensity_class": np.empty(0, dtype=np.uint8),
        }
    expansion = max(2, int(round(config.cell_expansion_um / mpp)))
    expanded = segmentation.expand_labels(nuclei.labels, distance=expansion)
    boundary = segmentation.find_boundaries(expanded, mode="inner") & (expanded > 0)
    labels = expanded[boundary].astype(np.int64)
    values = dab[boundary].astype(np.float64)
    pixels = np.bincount(labels, minlength=count + 1).astype(np.float64)
    sums = np.bincount(labels, weights=values, minlength=count + 1)
    positive = np.bincount(
        labels,
        weights=(values >= config.weak_dab_od).astype(np.float64),
        minlength=count + 1,
    )
    denominator = np.maximum(pixels[1:], 1.0)
    mean_values = sums[1:] / denominator
    coverage = positive[1:] / denominator
    return {
        "mean_membrane_dab_od": mean_values,
        "positive_membrane_fraction": coverage,
        "complete_membrane_proxy": coverage >= config.complete_membrane_fraction,
        "intensity_class": _intensity_classes(mean_values, config),
    }


def her2_pre_score(
    complete: np.ndarray,
    intensity: np.ndarray,
    coverage: np.ndarray,
    config: IHCConfig,
) -> int | None:
    """Return a conservative 0/1+/2+/3+ research pre-score."""
    count = int(len(intensity))
    if count < config.minimum_cells_for_score:
        return None
    strong_complete_pct = (
        100.0 * float(np.count_nonzero(complete & (intensity >= 3))) / count
    )
    weak_or_more_complete_pct = (
        100.0 * float(np.count_nonzero(complete & (intensity >= 1))) / count
    )
    incomplete_positive_pct = (
        100.0 * float(np.count_nonzero((~complete) & (coverage >= 0.10))) / count
    )
    threshold = config.her2_positive_cell_percent
    if strong_complete_pct > threshold:
        return 3
    if weak_or_more_complete_pct > threshold or 0 < strong_complete_pct <= threshold:
        return 2
    if incomplete_positive_pct > threshold:
        return 1
    return 0


PATCH_RESULT_FIELDS = (
    "schema_version",
    "engine_version",
    "analysis_signature",
    "completion_status",
    "case_alias",
    "patch_alias",
    "marker",
    "width",
    "height",
    "microns_per_pixel",
    "decoded_rgb_sha256",
    "decoded_rgb_verified",
    "tissue_fraction",
    "tissue_area_mm2",
    "nuclear_threshold_od",
    "cell_count",
    "cell_density_per_mm2",
    "dab_negative_cells",
    "dab_weak_cells",
    "dab_moderate_cells",
    "dab_strong_cells",
    "dab_positive_cells",
    "dab_positive_percent",
    "h_score",
    "complete_membrane_cells",
    "complete_membrane_percent",
    "strong_complete_membrane_cells",
    "strong_complete_membrane_percent",
    "incomplete_positive_membrane_cells",
    "incomplete_positive_membrane_percent",
    "her2_membrane_proxy_pre_score",
    "qc_status",
    "qc_flags",
    "cell_table",
    "qc_overlay",
    "error_type",
    "error_message",
)

CELL_FIELDS = (
    "cell_id",
    "centroid_x",
    "centroid_y",
    "area_um2",
    "mean_hematoxylin_od",
    "mean_dab_od",
    "dab_intensity_class",
    "dab_positive",
    "mean_membrane_dab_od",
    "positive_membrane_fraction",
    "complete_membrane_proxy",
)

CASE_RESULT_FIELDS = (
    "schema_version",
    "engine_version",
    "analysis_signature",
    "case_alias",
    "marker",
    "patches_scheduled",
    "patches_completed",
    "patches_failed",
    "tissue_area_mm2",
    "cell_count",
    "cell_density_per_mm2",
    "dab_negative_cells",
    "dab_weak_cells",
    "dab_moderate_cells",
    "dab_strong_cells",
    "dab_positive_cells",
    "dab_positive_percent",
    "h_score",
    "complete_membrane_cells",
    "complete_membrane_percent",
    "strong_complete_membrane_cells",
    "strong_complete_membrane_percent",
    "incomplete_positive_membrane_cells",
    "incomplete_positive_membrane_percent",
    "marker_pre_score",
    "marker_category",
    "qc_status",
    "qc_flags",
)


def _finite_or_blank(value: Any) -> Any:
    if isinstance(value, (float, np.floating)) and not math.isfinite(float(value)):
        return ""
    if isinstance(value, np.generic):
        return value.item()
    return value


def _number(value: Any, *, integer: bool = False) -> float | int | None:
    if value in (None, ""):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(parsed):
        return None
    return int(parsed) if integer else parsed


def _cell_rows_nuclear(
    nuclei: SegmentedNuclei, intensity: np.ndarray
) -> Iterable[dict[str, Any]]:
    for index in range(len(nuclei.label_ids)):
        yield {
            "cell_id": int(nuclei.label_ids[index]),
            "centroid_x": f"{nuclei.centroid_x[index]:.3f}",
            "centroid_y": f"{nuclei.centroid_y[index]:.3f}",
            "area_um2": f"{nuclei.area_um2[index]:.4f}",
            "mean_hematoxylin_od": f"{nuclei.mean_hematoxylin_od[index]:.6f}",
            "mean_dab_od": f"{nuclei.mean_dab_od[index]:.6f}",
            "dab_intensity_class": int(intensity[index]),
            "dab_positive": int(intensity[index] > 0),
            "mean_membrane_dab_od": "",
            "positive_membrane_fraction": "",
            "complete_membrane_proxy": "",
        }


def _cell_rows_her2(
    nuclei: SegmentedNuclei, membrane: Mapping[str, np.ndarray]
) -> Iterable[dict[str, Any]]:
    intensity = membrane["intensity_class"]
    for index in range(len(nuclei.label_ids)):
        yield {
            "cell_id": int(nuclei.label_ids[index]),
            "centroid_x": f"{nuclei.centroid_x[index]:.3f}",
            "centroid_y": f"{nuclei.centroid_y[index]:.3f}",
            "area_um2": f"{nuclei.area_um2[index]:.4f}",
            "mean_hematoxylin_od": f"{nuclei.mean_hematoxylin_od[index]:.6f}",
            "mean_dab_od": f"{nuclei.mean_dab_od[index]:.6f}",
            "dab_intensity_class": int(intensity[index]),
            "dab_positive": int(intensity[index] > 0),
            "mean_membrane_dab_od": (f"{membrane['mean_membrane_dab_od'][index]:.6f}"),
            "positive_membrane_fraction": (
                f"{membrane['positive_membrane_fraction'][index]:.6f}"
            ),
            "complete_membrane_proxy": int(membrane["complete_membrane_proxy"][index]),
        }


def _write_cell_table(path: Path, rows: Iterable[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    os.close(descriptor)
    try:
        with gzip.open(
            temporary, "wt", encoding="utf-8", newline="", compresslevel=6
        ) as handle:
            writer = csv.DictWriter(handle, fieldnames=list(CELL_FIELDS))
            writer.writeheader()
            writer.writerows(rows)
        os.replace(temporary, path)
    finally:
        try:
            os.unlink(temporary)
        except FileNotFoundError:
            pass


def _qc_colors() -> np.ndarray:
    return np.asarray(
        [
            [42, 82, 152],
            [241, 196, 15],
            [230, 126, 34],
            [185, 28, 28],
        ],
        dtype=np.uint8,
    )


def write_qc_overlay(
    path: Path,
    rgb: np.ndarray,
    nuclei: SegmentedNuclei,
    classes: np.ndarray,
    marker: str,
    label: str,
    maximum_edge: int = 1600,
) -> None:
    cv2, _tifffile, Image, ImageDraw, _ndimage, _feature, sk = (
        require_image_dependencies()
    )
    _measure, _morphology, segmentation = sk
    scale = min(1.0, maximum_edge / max(rgb.shape[:2]))
    width = max(1, int(round(rgb.shape[1] * scale)))
    height = max(1, int(round(rgb.shape[0] * scale)))
    base = cv2.resize(rgb, (width, height), interpolation=cv2.INTER_AREA)
    labels = cv2.resize(
        nuclei.labels.astype(np.int32),
        (width, height),
        interpolation=cv2.INTER_NEAREST,
    )
    boundaries = segmentation.find_boundaries(labels, mode="inner") & (labels > 0)
    class_lut = np.zeros(len(classes) + 1, dtype=np.uint8)
    if len(classes):
        class_lut[1:] = classes
    boundary_classes = class_lut[labels[boundaries]]
    colors = _qc_colors()
    overlay = base.copy()
    overlay[boundaries] = colors[boundary_classes]
    kernel = np.ones((2, 2), dtype=np.uint8)
    expanded = cv2.dilate(boundaries.astype(np.uint8), kernel, iterations=1).astype(
        bool
    )
    boundary_class_map = np.zeros(boundaries.shape, dtype=np.uint8)
    boundary_class_map[boundaries] = boundary_classes + 1
    nearest = cv2.dilate(boundary_class_map, kernel, iterations=1)
    expanded_classes = np.maximum(nearest[expanded].astype(np.int16) - 1, 0)
    overlay[expanded] = colors[expanded_classes]
    canvas = Image.fromarray(overlay, mode="RGB")
    draw = ImageDraw.Draw(canvas)
    draw.rectangle((0, 0, width, 34), fill=(255, 255, 255))
    draw.text(
        (10, 9),
        f"{marker} | {label} | cells={len(nuclei.label_ids):,}",
        fill=(20, 20, 20),
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".png", dir=path.parent
    )
    os.close(descriptor)
    try:
        canvas.save(temporary, format="PNG", optimize=True)
        os.replace(temporary, path)
    finally:
        try:
            os.unlink(temporary)
        except FileNotFoundError:
            pass


def quantify_patch(
    record: PatchRecord,
    config: IHCConfig,
    output_root: Path,
    *,
    verify_decoded_rgb: bool = True,
    save_cells: bool = False,
    save_qc: bool = True,
    resume: bool = True,
) -> dict[str, Any]:
    """Quantify one patch and write its independently resumable artifacts."""
    config.validate()
    signature = config.signature()
    patch_dir = output_root / record.output_relative
    result_path = patch_dir / "measurement.json"
    if resume and result_path.is_file():
        try:
            existing = json.loads(result_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            existing = {}
        if (
            existing.get("completion_status") == "completed"
            and existing.get("analysis_signature") == signature
            and existing.get("decoded_rgb_sha256") == record.expected_decoded_rgb_sha256
            and (not save_cells or (patch_dir / "cell_measurements.csv.gz").is_file())
            and (not save_qc or (patch_dir / "qc_overlay.png").is_file())
        ):
            existing["resume_reused"] = True
            return existing

    rgb = load_patch_rgb(record)
    decoded_hash = decoded_rgb_sha256(rgb)
    verified = decoded_hash == record.expected_decoded_rgb_sha256
    if verify_decoded_rgb and not verified:
        raise IHCError(
            f"Decoded RGB verification failed for {record.case_alias}/{record.patch_alias}"
        )
    hematoxylin, dab = separate_hematoxylin_dab(rgb)
    tissue = tissue_mask(rgb, record.microns_per_pixel)
    nuclei = segment_nuclei(
        hematoxylin, dab, tissue, record.marker, record.microns_per_pixel, config
    )
    tissue_fraction = float(np.mean(tissue))
    tissue_area_mm2 = (
        float(np.count_nonzero(tissue)) * record.microns_per_pixel**2 / 1_000_000.0
    )
    cell_count = int(len(nuclei.label_ids))
    cell_density = cell_count / tissue_area_mm2 if tissue_area_mm2 > 0 else math.nan
    qc_flags: list[str] = []
    if tissue_fraction < config.minimum_tissue_fraction:
        qc_flags.append("low_tissue_fraction")
    if cell_count < config.minimum_cells_for_score:
        qc_flags.append("low_cell_count")

    result: dict[str, Any] = {key: "" for key in PATCH_RESULT_FIELDS}
    result.update(
        {
            "schema_version": IHC_SCHEMA_VERSION,
            "engine_version": IHC_ENGINE_VERSION,
            "analysis_signature": signature,
            "completion_status": "completed",
            "case_alias": record.case_alias,
            "patch_alias": record.patch_alias,
            "marker": record.marker,
            "width": record.width,
            "height": record.height,
            "microns_per_pixel": record.microns_per_pixel,
            "decoded_rgb_sha256": decoded_hash,
            "decoded_rgb_verified": verified,
            "tissue_fraction": tissue_fraction,
            "tissue_area_mm2": tissue_area_mm2,
            "nuclear_threshold_od": nuclei.nuclear_threshold_od,
            "cell_count": cell_count,
            "cell_density_per_mm2": cell_density,
            "qc_status": "review" if qc_flags else "pass",
            "qc_flags": ";".join(qc_flags),
            "cell_table": "cell_measurements.csv.gz" if save_cells else "",
            "qc_overlay": "qc_overlay.png" if save_qc else "",
            "generated_at_utc": utc_now(),
            "research_use_only": True,
        }
    )

    if record.marker in NUCLEAR_MARKERS:
        intensity = _intensity_classes(nuclei.mean_dab_od, config)
        counts = np.bincount(intensity, minlength=4)
        positive = int(np.count_nonzero(intensity > 0))
        result.update(
            {
                "dab_negative_cells": int(counts[0]),
                "dab_weak_cells": int(counts[1]),
                "dab_moderate_cells": int(counts[2]),
                "dab_strong_cells": int(counts[3]),
                "dab_positive_cells": positive,
                "dab_positive_percent": (
                    100.0 * positive / cell_count if cell_count else math.nan
                ),
                "h_score": (
                    100.0 * float(np.sum(intensity)) / cell_count
                    if cell_count
                    else math.nan
                ),
            }
        )
        if save_cells:
            _write_cell_table(
                patch_dir / "cell_measurements.csv.gz",
                _cell_rows_nuclear(nuclei, intensity),
            )
        if save_qc:
            percent = _number(result["dab_positive_percent"])
            label = (
                f"DAB+ {percent:.1f}%" if percent is not None else "DAB+ unavailable"
            )
            write_qc_overlay(
                patch_dir / "qc_overlay.png",
                rgb,
                nuclei,
                intensity,
                record.marker,
                label,
            )
    else:
        membrane = her2_membrane_measurements(
            nuclei, dab, record.microns_per_pixel, config
        )
        intensity = membrane["intensity_class"]
        complete = membrane["complete_membrane_proxy"]
        coverage = membrane["positive_membrane_fraction"]
        counts = np.bincount(intensity, minlength=4)
        positive = int(np.count_nonzero(intensity > 0))
        complete_count = int(np.count_nonzero(complete & (intensity > 0)))
        strong_complete = int(np.count_nonzero(complete & (intensity >= 3)))
        incomplete_positive = int(np.count_nonzero((~complete) & (coverage >= 0.10)))
        pre_score = her2_pre_score(complete, intensity, coverage, config)
        result.update(
            {
                "dab_negative_cells": int(counts[0]),
                "dab_weak_cells": int(counts[1]),
                "dab_moderate_cells": int(counts[2]),
                "dab_strong_cells": int(counts[3]),
                "dab_positive_cells": positive,
                "dab_positive_percent": (
                    100.0 * positive / cell_count if cell_count else math.nan
                ),
                "complete_membrane_cells": complete_count,
                "complete_membrane_percent": (
                    100.0 * complete_count / cell_count if cell_count else math.nan
                ),
                "strong_complete_membrane_cells": strong_complete,
                "strong_complete_membrane_percent": (
                    100.0 * strong_complete / cell_count if cell_count else math.nan
                ),
                "incomplete_positive_membrane_cells": incomplete_positive,
                "incomplete_positive_membrane_percent": (
                    100.0 * incomplete_positive / cell_count if cell_count else math.nan
                ),
                "her2_membrane_proxy_pre_score": (
                    pre_score if pre_score is not None else ""
                ),
            }
        )
        if save_cells:
            _write_cell_table(
                patch_dir / "cell_measurements.csv.gz",
                _cell_rows_her2(nuclei, membrane),
            )
        if save_qc:
            label = (
                f"membrane-proxy pre-score {pre_score}+"
                if pre_score is not None
                else "pre-score unavailable"
            )
            write_qc_overlay(
                patch_dir / "qc_overlay.png",
                rgb,
                nuclei,
                intensity,
                record.marker,
                label,
            )

    portable = {key: _finite_or_blank(value) for key, value in result.items()}
    _atomic_json(result_path, portable)
    return portable


def _failed_patch_result(
    record: PatchRecord, config: IHCConfig, exc: BaseException
) -> dict[str, Any]:
    row = {key: "" for key in PATCH_RESULT_FIELDS}
    row.update(
        {
            "schema_version": IHC_SCHEMA_VERSION,
            "engine_version": IHC_ENGINE_VERSION,
            "analysis_signature": config.signature(),
            "completion_status": "failed",
            "case_alias": record.case_alias,
            "patch_alias": record.patch_alias,
            "marker": record.marker,
            "width": record.width,
            "height": record.height,
            "microns_per_pixel": record.microns_per_pixel,
            "decoded_rgb_verified": False,
            "qc_status": "failed",
            "qc_flags": "processing_failed",
            "error_type": type(exc).__name__,
            "error_message": " ".join(str(exc).split())[:500],
            "generated_at_utc": utc_now(),
            "research_use_only": True,
        }
    )
    return row


def _worker_quantify(payload: tuple[Any, ...]) -> dict[str, Any]:
    record, config, output_root, verify, save_cells, save_qc, resume = payload
    try:
        return quantify_patch(
            record,
            config,
            Path(output_root),
            verify_decoded_rgb=verify,
            save_cells=save_cells,
            save_qc=save_qc,
            resume=resume,
        )
    except BaseException as exc:
        return _failed_patch_result(record, config, exc)


def _her2_pre_score_from_counts(
    cell_count: int,
    complete_cells: int,
    strong_complete_cells: int,
    incomplete_positive_cells: int,
    config: IHCConfig,
) -> int | None:
    if cell_count < config.minimum_cells_for_score:
        return None
    threshold = config.her2_positive_cell_percent
    strong_percent = 100.0 * strong_complete_cells / cell_count
    complete_percent = 100.0 * complete_cells / cell_count
    incomplete_percent = 100.0 * incomplete_positive_cells / cell_count
    if strong_percent > threshold:
        return 3
    if complete_percent > threshold or 0 < strong_percent <= threshold:
        return 2
    if incomplete_percent > threshold:
        return 1
    return 0


def aggregate_case_results(
    patch_rows: Sequence[Mapping[str, Any]], config: IHCConfig
) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str], list[Mapping[str, Any]]] = defaultdict(list)
    for row in patch_rows:
        groups[(str(row["case_alias"]), str(row["marker"]))].append(row)
    output: list[dict[str, Any]] = []
    for (case_alias, marker), rows in sorted(
        groups.items(), key=lambda item: (item[0][0], IHC_MARKERS.index(item[0][1]))
    ):
        completed = [row for row in rows if row.get("completion_status") == "completed"]
        failed = [row for row in rows if row.get("completion_status") != "completed"]
        result = {field: "" for field in CASE_RESULT_FIELDS}
        result.update(
            {
                "schema_version": IHC_SCHEMA_VERSION,
                "engine_version": IHC_ENGINE_VERSION,
                "analysis_signature": config.signature(),
                "case_alias": case_alias,
                "marker": marker,
                "patches_scheduled": len(rows),
                "patches_completed": len(completed),
                "patches_failed": len(failed),
            }
        )
        if not completed:
            result.update({"qc_status": "failed", "qc_flags": "no_completed_patches"})
            output.append(result)
            continue
        sums: dict[str, float] = {}
        for field in (
            "tissue_area_mm2",
            "cell_count",
            "dab_negative_cells",
            "dab_weak_cells",
            "dab_moderate_cells",
            "dab_strong_cells",
            "dab_positive_cells",
            "complete_membrane_cells",
            "strong_complete_membrane_cells",
            "incomplete_positive_membrane_cells",
        ):
            values = [_number(row.get(field)) for row in completed]
            sums[field] = sum(value for value in values if value is not None)
        cells = int(sums["cell_count"])
        tissue_area = sums["tissue_area_mm2"]
        positive = int(sums["dab_positive_cells"])
        result.update(
            {
                "tissue_area_mm2": tissue_area,
                "cell_count": cells,
                "cell_density_per_mm2": (
                    cells / tissue_area if tissue_area > 0 else ""
                ),
                "dab_negative_cells": int(sums["dab_negative_cells"]),
                "dab_weak_cells": int(sums["dab_weak_cells"]),
                "dab_moderate_cells": int(sums["dab_moderate_cells"]),
                "dab_strong_cells": int(sums["dab_strong_cells"]),
                "dab_positive_cells": positive,
                "dab_positive_percent": 100.0 * positive / cells if cells else "",
            }
        )
        flags = sorted(
            {
                flag
                for row in completed
                for flag in str(row.get("qc_flags", "")).split(";")
                if flag
            }
        )
        if failed:
            flags.append("one_or_more_failed_patches")
        if marker in NUCLEAR_MARKERS:
            weighted = (
                sums["dab_weak_cells"]
                + 2 * sums["dab_moderate_cells"]
                + 3 * sums["dab_strong_cells"]
            )
            result["h_score"] = 100.0 * weighted / cells if cells else ""
            percent = _number(result["dab_positive_percent"])
            if percent is not None:
                result["marker_pre_score"] = percent
                if marker in {"ER", "PR"}:
                    result["marker_category"] = (
                        "positive"
                        if percent >= config.er_pr_positive_percent
                        else "negative"
                    )
                else:
                    result["marker_category"] = (
                        "high-at-20-percent"
                        if percent >= config.ki67_high_percent
                        else "low-at-20-percent"
                    )
        else:
            complete = int(sums["complete_membrane_cells"])
            strong = int(sums["strong_complete_membrane_cells"])
            incomplete = int(sums["incomplete_positive_membrane_cells"])
            result.update(
                {
                    "complete_membrane_cells": complete,
                    "complete_membrane_percent": (
                        100.0 * complete / cells if cells else ""
                    ),
                    "strong_complete_membrane_cells": strong,
                    "strong_complete_membrane_percent": (
                        100.0 * strong / cells if cells else ""
                    ),
                    "incomplete_positive_membrane_cells": incomplete,
                    "incomplete_positive_membrane_percent": (
                        100.0 * incomplete / cells if cells else ""
                    ),
                }
            )
            score = _her2_pre_score_from_counts(
                cells, complete, strong, incomplete, config
            )
            if score is not None:
                result["marker_pre_score"] = score
                result["marker_category"] = f"{score}+"
        result["qc_status"] = "review" if flags else "pass"
        result["qc_flags"] = ";".join(sorted(set(flags)))
        output.append({key: _finite_or_blank(value) for key, value in result.items()})
    return output


def _format_number(value: Any, digits: int = 1) -> str:
    parsed = _number(value)
    return "—" if parsed is None else f"{parsed:,.{digits}f}"


def _marker_measurement_label(row: Mapping[str, Any]) -> str:
    if row.get("marker") == "HER2":
        score = row.get("marker_pre_score", "")
        return f"{score}+ membrane-proxy pre-score" if score != "" else "unavailable"
    percent = _number(row.get("dab_positive_percent"))
    return f"{percent:.1f}% DAB+" if percent is not None else "unavailable"


def write_case_ihc_reports(
    output_root: Path,
    patch_rows: Sequence[Mapping[str, Any]],
    case_rows: Sequence[Mapping[str, Any]],
) -> dict[str, str]:
    patches_by_case: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    measurements_by_case: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in patch_rows:
        patches_by_case[str(row["case_alias"])].append(row)
    for row in case_rows:
        measurements_by_case[str(row["case_alias"])].append(row)

    links: dict[str, str] = {}
    for case_alias in sorted(measurements_by_case):
        marker_cards: list[str] = []
        for row in sorted(
            measurements_by_case[case_alias],
            key=lambda value: IHC_MARKERS.index(str(value["marker"])),
        ):
            marker_cards.append(
                "<article class='metric'>"
                f"<span>{html.escape(str(row['marker']))}</span>"
                f"<b>{html.escape(_marker_measurement_label(row))}</b>"
                f"<small>{int(_number(row.get('cell_count'), integer=True) or 0):,} "
                f"objects · QC {html.escape(str(row.get('qc_status', '')))}</small>"
                "</article>"
            )

        patch_cards: list[str] = []
        for row in sorted(
            patches_by_case.get(case_alias, []),
            key=lambda value: (
                IHC_MARKERS.index(str(value["marker"])),
                str(value["patch_alias"]),
            ),
        ):
            marker = str(row["marker"])
            patch_alias = str(row["patch_alias"])
            value = _marker_measurement_label(
                {
                    "marker": marker,
                    "marker_pre_score": row.get("her2_membrane_proxy_pre_score", ""),
                    "dab_positive_percent": row.get("dab_positive_percent", ""),
                }
            )
            artifact_root = output_root / "patches" / case_alias / patch_alias
            overlay = artifact_root / "qc_overlay.png"
            cells = artifact_root / "cell_measurements.csv.gz"
            relative_root = f"../patches/{case_alias}/{patch_alias}"
            image_html = (
                f"<a href='{relative_root}/qc_overlay.png'>"
                f"<img loading='lazy' src='{relative_root}/qc_overlay.png' "
                f"alt='QC overlay for {html.escape(patch_alias)}'></a>"
                if overlay.is_file()
                else "<div class='noimage'>QC overlay not written</div>"
            )
            cell_link = (
                f" · <a href='{relative_root}/cell_measurements.csv.gz'>cells</a>"
                if cells.is_file()
                else ""
            )
            patch_cards.append(
                "<article class='patch'>"
                f"{image_html}<div><b>{html.escape(marker)} · "
                f"{html.escape(value)}</b><code>{html.escape(patch_alias)}</code>"
                f"<small>{int(_number(row.get('cell_count'), integer=True) or 0):,} "
                f"objects · QC {html.escape(str(row.get('qc_status', '')))}"
                f"{cell_link}</small></div></article>"
            )

        page = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{html.escape(case_alias)} · TumorQuantAI IHC QC</title>
<style>
:root{{--ink:#172033;--muted:#5f6b7a;--navy:#18324a;--blue:#2a6f97;--paper:#f5f7fa;--line:#dbe2ea}}
*{{box-sizing:border-box}}body{{margin:0;background:var(--paper);color:var(--ink);font:15px/1.5 system-ui,sans-serif}}
header{{background:linear-gradient(135deg,var(--navy),#285d78);color:white;padding:34px max(20px,calc((100% - 1180px)/2))}}
main{{max-width:1180px;margin:auto;padding:24px}}a{{color:var(--blue)}}h1{{margin:5px 0;font-size:26px}}header a{{color:#d8edf7}}
.metrics,.gallery{{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:14px}}
.metric,.patch{{background:white;border:1px solid var(--line);border-radius:13px;overflow:hidden;box-shadow:0 8px 24px #24384c12}}
.metric{{padding:18px}}.metric span,.metric small,.patch small,.patch code{{display:block;color:var(--muted)}}.metric b{{display:block;font-size:20px;margin:4px 0}}
.patch img,.noimage{{width:100%;aspect-ratio:3/2;object-fit:cover;background:#e9eef3}}.noimage{{display:grid;place-items:center;color:var(--muted)}}
.patch>div{{padding:12px}}.patch code{{font-size:10px;margin:5px 0;overflow-wrap:anywhere}}h2{{margin:30px 0 12px}}
.notice{{background:#fff8e8;border-left:5px solid #e4a11b;padding:12px 16px;margin:20px 0}}
</style></head><body><header><a href="../START_HERE.html">← cohort report</a><h1>{html.escape(case_alias)}</h1><div>Marker measurements and patch-level segmentation QC · research use only</div></header>
<main><div class="notice">Selected fields are not whole-slide measurements. No verified invasive-tumour ROI is supplied; inspect every overlay before interpreting a research pre-score.</div>
<section class="metrics">{''.join(marker_cards)}</section><h2>Patch QC gallery</h2><section class="gallery">{''.join(patch_cards)}</section></main></body></html>"""
        relative_page = f"case_reports/{case_alias}.html"
        _atomic_text(output_root / relative_page, page)
        links[case_alias] = relative_page
    return links


def write_ihc_report(
    output_root: Path,
    patch_rows: Sequence[Mapping[str, Any]],
    case_rows: Sequence[Mapping[str, Any]],
    run_manifest: Mapping[str, Any],
) -> Path:
    completed = sum(row.get("completion_status") == "completed" for row in patch_rows)
    failed = len(patch_rows) - completed
    cases = sorted({str(row["case_alias"]) for row in case_rows})
    case_report_links = write_case_ihc_reports(output_root, patch_rows, case_rows)
    total_cells = sum(
        int(_number(row.get("cell_count"), integer=True) or 0) for row in case_rows
    )
    cohort_cards: list[str] = []
    for marker in IHC_MARKERS:
        rows = [row for row in case_rows if row.get("marker") == marker]
        qc_counts = Counter(str(row.get("qc_status", "unknown")) for row in rows)
        qc_text = " · ".join(
            f"QC {status}: {count}" for status, count in sorted(qc_counts.items())
        )
        if marker == "HER2":
            scores = [
                int(value)
                for row in rows
                if (value := _number(row.get("marker_pre_score"), integer=True))
                is not None
            ]
            counts = Counter(scores)
            headline = f"{len(scores)} scored cases"
            detail = " · ".join(f"{score}+: {counts[score]}" for score in range(4))
        else:
            values = [
                value
                for row in rows
                if (value := _number(row.get("dab_positive_percent"))) is not None
            ]
            if values:
                cutoff = 20.0 if marker == "Ki-67" else 1.0
                headline = f"median {float(np.median(values)):.1f}% DAB+"
                detail = (
                    f"range {min(values):.1f}–{max(values):.1f}% · "
                    f"≥{cutoff:g}%: {sum(value >= cutoff for value in values)}/{len(values)}"
                )
            else:
                headline = "unavailable"
                detail = "No scored cases"
        cohort_cards.append(
            "<article class='marker'>"
            f"<span>{html.escape(marker)}</span><b>{html.escape(headline)}</b>"
            f"<small>{html.escape(detail)}<br>{html.escape(qc_text)}</small>"
            "</article>"
        )
    marker_rows = []
    for row in case_rows:
        case_alias = str(row["case_alias"])
        if row["marker"] == "HER2":
            value = (
                f"{row['marker_pre_score']}+ pre-score"
                if row.get("marker_pre_score") != ""
                else "unavailable"
            )
        else:
            value = _format_number(row.get("dab_positive_percent")) + "% DAB+"
        marker_rows.append(
            "<tr>"
            f"<td><a href='{html.escape(case_report_links[case_alias])}'>"
            f"<code>{html.escape(case_alias)}</code></a></td>"
            f"<td>{html.escape(str(row['marker']))}</td>"
            f"<td>{html.escape(value)}</td>"
            f"<td>{int(_number(row.get('cell_count'), integer=True) or 0):,}</td>"
            f"<td><span class='status {html.escape(str(row['qc_status']))}'>"
            f"{html.escape(str(row['qc_status']))}</span></td></tr>"
        )
    html_text = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>TumorQuantAI IHC report</title>
<style>
:root{{--ink:#172033;--muted:#5f6b7a;--navy:#18324a;--blue:#2a6f97;--gold:#e4a11b;--paper:#f5f7fa;--line:#dbe2ea;--danger:#9f2d2d}}
*{{box-sizing:border-box}} body{{margin:0;background:var(--paper);color:var(--ink);font:15px/1.55 system-ui,-apple-system,Segoe UI,sans-serif}}
header{{background:linear-gradient(135deg,var(--navy),#285d78);color:white;padding:48px max(24px,calc((100% - 1180px)/2)) 42px}}
header p{{max-width:780px;color:#dceaf2;margin:8px 0 0}} main{{max-width:1180px;margin:0 auto;padding:26px 22px 60px}}
.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:14px;margin-top:-48px}}
.card,.panel{{background:white;border:1px solid var(--line);border-radius:14px;box-shadow:0 8px 26px #24384c12}}
.card{{padding:20px}} .card b{{display:block;font-size:28px;color:var(--navy)}} .card span{{color:var(--muted)}}
.panel{{padding:24px;margin-top:22px}} .warning{{border-left:5px solid var(--gold);background:#fffaf0}}
.marker-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(225px,1fr));gap:14px;margin-top:14px}}
.marker{{border:1px solid var(--line);border-radius:11px;padding:16px;background:#f8fafc}}.marker span,.marker small{{display:block;color:var(--muted)}}.marker b{{display:block;font-size:18px;color:var(--navy);margin:3px 0}}
h1,h2{{margin:0 0 8px}} h2{{font-size:20px}} a{{color:var(--blue)}}
table{{border-collapse:collapse;width:100%;font-size:14px}} th,td{{border-bottom:1px solid var(--line);padding:9px;text-align:left}} th{{position:sticky;top:0;background:#f8fafc}}
.scroll{{max-height:620px;overflow:auto;border:1px solid var(--line);border-radius:10px}}
.status{{padding:3px 8px;border-radius:999px;background:#e9f5ee;color:#25633b}} .status.review{{background:#fff2d8;color:#7d5100}} .status.failed{{background:#fde9e9;color:var(--danger)}} code{{font-size:12px}}
</style></head><body>
<header><h1>TumorQuantAI IHC quantification</h1><p>Deterministic H–DAB stain separation, physical-scale-aware nuclear segmentation, marker measurements, and reviewable QC. Research use only.</p></header>
<main><section class="cards"><div class="card"><b>{len(cases)}</b><span>cases</span></div><div class="card"><b>{completed:,}</b><span>completed patches</span></div><div class="card"><b>{failed:,}</b><span>failed patches</span></div><div class="card"><b>{total_cells:,}</b><span>segmented measurement proxies</span></div></section>
<section class="panel warning"><h2>Interpretation boundary</h2><p>These selected fields are not whole-slide measurements. No pathologist-verified invasive-tumour ROI or independently validated tumour-cell classifier is supplied. ER/PR/Ki-67 values include every accepted segmented nucleus; HER2 is an expanded-nucleus boundary proxy, not a clinical membrane score. Separate stains are not registered and do not measure same-cell co-expression. Review every overlay and do not use these outputs for patient care.</p></section>
<section class="panel"><h2>Open the audit tables first</h2><p><a href="tables/patch_measurements.csv">Patch measurements</a> · <a href="tables/case_marker_measurements.csv">Case-marker measurements</a> · <a href="workflow_metadata/ihc_run.json">Run provenance</a></p></section>
<section class="panel"><h2>Cohort overview</h2><p>Descriptive research summaries across available case-level measurements.</p><div class="marker-grid">{''.join(cohort_cards)}</div></section>
<section class="panel"><h2>Case-marker summary</h2><div class="scroll"><table><thead><tr><th>Public case alias</th><th>Marker</th><th>Measurement</th><th>Cells</th><th>QC</th></tr></thead><tbody>{''.join(marker_rows)}</tbody></table></div></section>
<section class="panel"><h2>Method identity and sources</h2><p><code>{html.escape(str(run_manifest.get('engine_version', IHC_ENGINE_VERSION)))}</code><br>Analysis signature: <code>{html.escape(str(run_manifest.get('analysis_signature', '')))}</code><br>Dataset DOI: <a href="https://doi.org/{PUBLIC_DATASET_DOI}">{PUBLIC_DATASET_DOI}</a></p><ul><li><a href="https://pubmed.ncbi.nlm.nih.gov/11531144/">Ruifrok &amp; Johnston (2001)</a>: H–DAB colour deconvolution foundation.</li><li><a href="https://ascopubs.org/doi/10.1200/JCO.19.02309">ASCO/CAP ER and PR guideline update</a>.</li><li><a href="https://www.cap.org/cap-guidelines/her2-testing-in-breast-cancer-2023-guideline-update/">CAP HER2 testing guideline update</a>.</li><li><a href="https://pmc.ncbi.nlm.nih.gov/articles/PMC8487652/">International Ki67 Working Group recommendations</a>.</li></ul></section>
</main></body></html>"""
    path = output_root / "START_HERE.html"
    _atomic_text(path, html_text)
    return path


def run_quantification(
    records: Sequence[PatchRecord],
    unavailable: Sequence[Mapping[str, str]],
    manifest_path: Path,
    input_root: Path,
    output_root: Path,
    config: IHCConfig,
    *,
    workers: int,
    verify_decoded_rgb: bool,
    save_cells: bool,
    save_qc: bool,
    resume: bool,
    allow_missing: bool,
    fail_fast: bool,
    dry_run: bool,
) -> dict[str, Any]:
    config.validate()
    if not records:
        raise IHCError("No available IHC patches matched the requested selection")
    if unavailable and not allow_missing:
        raise IHCError(
            f"{len(unavailable)} selected patches are unavailable; complete the archive "
            "download or pass --allow-missing for an explicitly incomplete run"
        )
    output_root = output_root.expanduser().resolve()
    input_root = input_root.expanduser().resolve()
    if (
        output_root == input_root
        or output_root in input_root.parents
        or input_root in output_root.parents
    ):
        raise IHCError("IHC input and output directories must not overlap")
    manifest_path = manifest_path.expanduser().resolve()
    manifest_digest = hashlib.sha256(manifest_path.read_bytes()).hexdigest()
    run_path = output_root / "workflow_metadata/ihc_run.json"
    existing: dict[str, Any] = {}
    if run_path.is_file():
        try:
            existing = json.loads(run_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise IHCError("Existing IHC run manifest is unreadable") from exc
        if existing.get("analysis_signature") != config.signature():
            raise IHCError(
                "Output contains a different IHC analysis signature; choose a new output"
            )
        if existing.get("source_manifest_sha256") != manifest_digest:
            raise IHCError(
                "Output belongs to a different patch manifest; choose a new output"
            )
    elif output_root.exists() and any(output_root.iterdir()):
        raise IHCError("Output is non-empty but has no TumorQuantAI IHC run manifest")
    workers = max(1, min(int(workers), len(records)))
    plan = {
        "schema_version": IHC_SCHEMA_VERSION,
        "engine_version": IHC_ENGINE_VERSION,
        "analysis_signature": config.signature(),
        "source_manifest_sha256": manifest_digest,
        "dataset_record": PUBLIC_DATASET_RECORD,
        "dataset_doi": PUBLIC_DATASET_DOI,
        "selected_patch_count": len(records),
        "selected_case_count": len({record.case_alias for record in records}),
        "unavailable_patch_count": len(unavailable),
        "markers": sorted({record.marker for record in records}, key=IHC_MARKERS.index),
        "workers": workers,
        "verify_decoded_rgb": verify_decoded_rgb,
        "save_cell_tables": save_cells,
        "save_qc_overlays": save_qc,
        "settings": asdict(config),
        "research_use_only": True,
    }
    if dry_run:
        return {**plan, "status": "dry_run"}
    output_root.mkdir(parents=True, exist_ok=True)
    running = {
        **plan,
        "status": "running",
        "started_at_utc": existing.get("started_at_utc", utc_now()),
        "updated_at_utc": utc_now(),
    }
    _atomic_json(run_path, running)
    payloads = [
        (
            record,
            config,
            str(output_root),
            verify_decoded_rgb,
            save_cells,
            save_qc,
            resume,
        )
        for record in records
    ]
    results: list[dict[str, Any]] = []
    if workers == 1:
        for index, row in enumerate(map(_worker_quantify, payloads), start=1):
            results.append(row)
            print(
                f"IHC {index}/{len(records)} {row['case_alias']} "
                f"{row['marker']} {row['completion_status']}",
                flush=True,
            )
            if fail_fast and row["completion_status"] != "completed":
                break
    else:
        with ProcessPoolExecutor(max_workers=workers) as executor:
            future_rows = {
                executor.submit(_worker_quantify, payload): payload[0]
                for payload in payloads
            }
            for index, future in enumerate(as_completed(future_rows), start=1):
                record = future_rows[future]
                try:
                    row = future.result()
                except BaseException as exc:
                    row = _failed_patch_result(record, config, exc)
                results.append(row)
                print(
                    f"IHC {index}/{len(records)} {row['case_alias']} "
                    f"{row['marker']} {row['completion_status']}",
                    flush=True,
                )
                if fail_fast and row["completion_status"] != "completed":
                    for pending in future_rows:
                        pending.cancel()
                    break
    completed_keys = {(row["case_alias"], row["patch_alias"]) for row in results}
    if len(completed_keys) < len(records):
        for record in records:
            if (record.case_alias, record.patch_alias) not in completed_keys:
                results.append(
                    _failed_patch_result(
                        record, config, IHCError("not processed after fail-fast")
                    )
                )
    results.sort(
        key=lambda row: (
            str(row["case_alias"]),
            IHC_MARKERS.index(str(row["marker"])),
            str(row["patch_alias"]),
        )
    )
    for row in results:
        if row["completion_status"] != "completed":
            patch_dir = (
                output_root
                / "patches"
                / str(row["case_alias"])
                / str(row["patch_alias"])
            )
            _atomic_json(patch_dir / "measurement.json", row)
    case_rows = aggregate_case_results(results, config)
    _write_csv(
        output_root / "tables/patch_measurements.csv",
        results,
        (*PATCH_RESULT_FIELDS, "generated_at_utc", "research_use_only"),
    )
    _write_csv(
        output_root / "tables/case_marker_measurements.csv",
        case_rows,
        CASE_RESULT_FIELDS,
    )
    _write_csv(
        output_root / "tables/unavailable_patches.csv",
        list(unavailable),
        ("case_alias", "patch_alias", "marker", "public_path"),
    )
    failed_count = sum(row["completion_status"] != "completed" for row in results)
    final = {
        **running,
        "status": (
            "completed_with_failures" if failed_count or unavailable else "completed"
        ),
        "completed_patch_count": len(results) - failed_count,
        "failed_patch_count": failed_count,
        "completed_case_marker_count": sum(
            int(row["patches_completed"]) > 0 for row in case_rows
        ),
        "finished_at_utc": utc_now(),
        "updated_at_utc": utc_now(),
        "outputs": {
            "report": "START_HERE.html",
            "case_qc_reports": "case_reports/",
            "patch_measurements": "tables/patch_measurements.csv",
            "case_marker_measurements": "tables/case_marker_measurements.csv",
            "unavailable_patches": "tables/unavailable_patches.csv",
        },
    }
    _atomic_json(run_path, final)
    report = write_ihc_report(output_root, results, case_rows, final)
    return {**final, "report_path": str(report)}


def _normalize_identifier(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).casefold()
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)) and math.isfinite(float(value)):
        number = float(value)
        return str(int(number)) if number.is_integer() else format(number, ".15g")
    return " ".join(str(value).strip().casefold().split())


def _is_public_case_alias(value: str) -> bool:
    prefix = "TQA_BC_"
    suffix = value[len(prefix) :] if value.startswith(prefix) else ""
    return len(suffix) == 20 and set(suffix) <= set("ABCDEFGHIJKLMNOPQRSTUVWXYZ234567")


PATHOLOGIST_FIELDS = (
    "case_alias",
    "pathologist_er_percent",
    "pathologist_pr_percent",
    "pathologist_her2_ihc_score",
    "pathologist_her2_fish",
    "pathologist_ki67_percent",
)

DEFAULT_CLINICAL_COLUMNS = {
    "pathologist_er_percent": "Receptor de estrógeno (%)",
    "pathologist_pr_percent": "Receptor de progesterona (%)",
    "pathologist_her2_ihc_score": "HER-2 (IHQ)",
    "pathologist_her2_fish": (
        "Amplificación FISH HER2 (No realizado=0, no amplificado=1, amplificado=2)"
    ),
    "pathologist_ki67_percent": "Ki67",
}


def _load_linkage_rows(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    if not path.is_file():
        raise IHCError(f"Private case linkage does not exist: {path}")
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = [dict(row) for row in reader]
        fields = list(reader.fieldnames or [])
    if "case_alias" not in fields:
        raise IHCError("Private linkage must contain case_alias")
    invalid_aliases = [
        str(row.get("case_alias", "")).strip()
        for row in rows
        if not _is_public_case_alias(str(row.get("case_alias", "")).strip())
    ]
    if invalid_aliases:
        raise IHCError("Private linkage contains a non-public case_alias value")
    return rows, fields


def export_pathologist_csv(
    workbook: Path,
    linkage: Path,
    output_csv: Path,
    *,
    sheet_name: str = "Biopsias finales incluidas",
    clinical_id_column: str | None = None,
    linkage_id_column: str | None = None,
    clinical_columns: Mapping[str, str] | None = None,
) -> dict[str, Any]:
    """Create a privacy-minimized, public-alias keyed marker table."""
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise IHCError(
            "openpyxl is required for the pathologist XLSX export; rerun the installer"
        ) from exc
    workbook = workbook.expanduser().resolve()
    linkage = linkage.expanduser().resolve()
    output_csv = output_csv.expanduser().resolve()
    if not workbook.is_file():
        raise IHCError(f"Pathologist workbook does not exist: {workbook}")
    if output_csv.exists():
        raise IHCError(f"Refusing to overwrite existing clinical export: {output_csv}")
    book = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    if sheet_name not in book.sheetnames:
        raise IHCError(f"Workbook sheet does not exist: {sheet_name}")
    sheet = book[sheet_name]
    values = sheet.iter_rows(values_only=True)
    try:
        headers = [
            str(value).strip() if value is not None else "" for value in next(values)
        ]
    except StopIteration as exc:
        raise IHCError("Pathologist workbook sheet is empty") from exc
    if len(headers) != len(set(headers)):
        raise IHCError("Pathologist sheet contains duplicate column headings")
    clinical_rows = [
        dict(zip(headers, row))
        for row in values
        if any(value is not None for value in row)
    ]
    linkage_rows, linkage_fields = _load_linkage_rows(linkage)
    clinical_candidates = (
        [clinical_id_column]
        if clinical_id_column
        else [
            name
            for name in ("Número de paciente", "Biopsia", "case_id")
            if name in headers
        ]
    )
    linkage_candidates = (
        [linkage_id_column]
        if linkage_id_column
        else [
            name
            for name in ("case_id", "clinical_case_id", "patient_number")
            if name in linkage_fields
        ]
    )
    if not clinical_candidates or not linkage_candidates:
        raise IHCError(
            "Could not identify linkage columns; pass --clinical-id-column and "
            "--linkage-id-column explicitly"
        )
    matches: list[tuple[str, str, dict[str, str], dict[str, dict[str, Any]]]] = []
    for clinical_column in clinical_candidates:
        if clinical_column not in headers:
            continue
        clinical_by_id: dict[str, dict[str, Any]] = {}
        duplicate = False
        for row in clinical_rows:
            key = _normalize_identifier(row.get(clinical_column))
            if not key or key in clinical_by_id:
                duplicate = True
                break
            clinical_by_id[key] = row
        if duplicate:
            continue
        for linkage_column in linkage_candidates:
            if linkage_column not in linkage_fields:
                continue
            alias_to_id: dict[str, str] = {}
            consistent = True
            for row in linkage_rows:
                alias = str(row.get("case_alias", "")).strip()
                key = _normalize_identifier(row.get(linkage_column))
                if not alias or not key:
                    consistent = False
                    break
                prior = alias_to_id.setdefault(alias, key)
                if prior != key:
                    consistent = False
                    break
            if (
                consistent
                and alias_to_id
                and all(key in clinical_by_id for key in alias_to_id.values())
                and len(set(alias_to_id.values())) == len(alias_to_id)
            ):
                matches.append(
                    (clinical_column, linkage_column, alias_to_id, clinical_by_id)
                )
    if len(matches) != 1:
        raise IHCError(
            "Private linkage could not be matched uniquely to the workbook. "
            "Specify both identifier columns; do not infer linkage from marker values."
        )
    clinical_column, linkage_column, alias_to_id, clinical_by_id = matches[0]
    columns = dict(DEFAULT_CLINICAL_COLUMNS)
    if clinical_columns:
        columns.update(clinical_columns)
    missing_columns = [column for column in columns.values() if column not in headers]
    if missing_columns:
        raise IHCError(
            "Pathologist sheet is missing marker columns: " + ", ".join(missing_columns)
        )
    exported: list[dict[str, Any]] = []
    for alias, identifier in sorted(alias_to_id.items()):
        source = clinical_by_id[identifier]
        row: dict[str, Any] = {"case_alias": alias}
        for output_field, source_field in columns.items():
            value = _number(source.get(source_field))
            if value is None:
                raise IHCError(
                    f"Missing or non-numeric {source_field!r} for public alias {alias}"
                )
            row[output_field] = value
        for field in (
            "pathologist_er_percent",
            "pathologist_pr_percent",
            "pathologist_ki67_percent",
        ):
            if not 0 <= float(row[field]) <= 100:
                raise IHCError(f"Clinical percentage outside [0, 100] for {alias}")
        if float(row["pathologist_her2_ihc_score"]) not in {0.0, 1.0, 2.0, 3.0}:
            raise IHCError(f"HER2 IHC score outside 0-3 for {alias}")
        if float(row["pathologist_her2_fish"]) not in {0.0, 1.0, 2.0}:
            raise IHCError(f"HER2 FISH code outside 0-2 for {alias}")
        exported.append(row)
    _write_csv(output_csv, exported, PATHOLOGIST_FIELDS)
    provenance = {
        "schema_version": "tumorquantai_pathologist_export_v1",
        "privacy_status": "pseudonymized_minimum_marker_table",
        "rows": len(exported),
        "included_columns": list(PATHOLOGIST_FIELDS),
        "excluded_data": (
            "names, RUT/national identifiers, biopsy identifiers, dates, age, "
            "diagnosis text, laterality, specimen type, and grade"
        ),
        "clinical_sheet": sheet_name,
        "clinical_link_column": clinical_column,
        "private_linkage_column": linkage_column,
        "source_workbook_sha256": hashlib.sha256(workbook.read_bytes()).hexdigest(),
        "generated_at_utc": utc_now(),
        "warning": (
            "Public aliases plus clinical marker values remain pseudonymized data; "
            "protect this export and never publish the private linkage."
        ),
    }
    _atomic_json(
        output_csv.with_suffix(output_csv.suffix + ".provenance.json"), provenance
    )
    return {**provenance, "output_csv": str(output_csv)}


def cohen_kappa(
    reference: Sequence[int],
    prediction: Sequence[int],
    categories: Sequence[int],
    *,
    weights: str = "none",
) -> tuple[float, np.ndarray]:
    if len(reference) != len(prediction) or not reference:
        return math.nan, np.zeros((len(categories), len(categories)), dtype=np.int64)
    category_index = {value: index for index, value in enumerate(categories)}
    confusion = np.zeros((len(categories), len(categories)), dtype=np.int64)
    for truth, predicted in zip(reference, prediction):
        if truth not in category_index or predicted not in category_index:
            raise IHCError("Kappa category is outside the prespecified scale")
        confusion[category_index[truth], category_index[predicted]] += 1
    total = float(confusion.sum())
    observed = confusion / total
    expected = np.outer(confusion.sum(axis=1), confusion.sum(axis=0)) / (total * total)
    size = len(categories)
    if weights == "none":
        disagreement = np.ones((size, size), dtype=np.float64) - np.eye(size)
    elif weights in {"linear", "quadratic"}:
        indices = np.arange(size, dtype=np.float64)
        disagreement = np.abs(indices[:, None] - indices[None, :])
        if size > 1:
            disagreement /= size - 1
        if weights == "quadratic":
            disagreement **= 2
    else:
        raise IHCError(f"Unknown kappa weighting: {weights}")
    observed_disagreement = float(np.sum(disagreement * observed))
    expected_disagreement = float(np.sum(disagreement * expected))
    if expected_disagreement <= np.finfo(float).eps:
        value = 1.0 if observed_disagreement <= np.finfo(float).eps else math.nan
    else:
        value = 1.0 - observed_disagreement / expected_disagreement
    return value, confusion


def bootstrap_kappa_interval(
    reference: Sequence[int],
    prediction: Sequence[int],
    categories: Sequence[int],
    *,
    weights: str,
    iterations: int,
    seed: int,
) -> tuple[float, float]:
    if iterations <= 0 or len(reference) < 2:
        return math.nan, math.nan
    truth = np.asarray(reference, dtype=np.int64)
    predicted = np.asarray(prediction, dtype=np.int64)
    generator = np.random.default_rng(seed)
    values: list[float] = []
    for _ in range(iterations):
        indices = generator.integers(0, len(truth), size=len(truth))
        value, _ = cohen_kappa(
            truth[indices].tolist(),
            predicted[indices].tolist(),
            categories,
            weights=weights,
        )
        if math.isfinite(value):
            values.append(value)
    if len(values) < max(20, iterations // 10):
        return math.nan, math.nan
    low, high = np.percentile(values, [2.5, 97.5])
    return float(low), float(high)


def _percentage_decile(value: float) -> int:
    return min(9, max(0, int(float(value) // 10)))


def compare_pathologist_agreement(
    results_root: Path,
    pathologist_csv: Path,
    output_dir: Path,
    *,
    bootstrap_iterations: int = 2000,
    bootstrap_seed: int = 20260829,
) -> dict[str, Any]:
    case_path = results_root / "tables/case_marker_measurements.csv"
    if not case_path.is_file():
        raise IHCError(f"Case-marker measurements do not exist: {case_path}")
    if not pathologist_csv.is_file():
        raise IHCError(f"Pathologist CSV does not exist: {pathologist_csv}")
    with case_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if {"case_alias", "marker", "marker_pre_score"} - set(reader.fieldnames or []):
            raise IHCError(
                "Case-marker measurements do not use the required IHC schema"
            )
        algorithm_rows = list(reader)
    with pathologist_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if set(PATHOLOGIST_FIELDS) - set(reader.fieldnames or []):
            raise IHCError(
                "Pathologist CSV does not use the required privacy-minimized schema"
            )
        pathologist_rows = list(reader)
    if not pathologist_rows:
        raise IHCError("Pathologist CSV contains no data rows")
    if any(
        not _is_public_case_alias(str(row.get("case_alias", "")).strip())
        for row in pathologist_rows
    ):
        raise IHCError("Pathologist CSV contains a non-public case_alias value")
    clinical_by_alias = {row["case_alias"]: row for row in pathologist_rows}
    if len(clinical_by_alias) != len(pathologist_rows):
        raise IHCError("Pathologist CSV contains duplicate case_alias values")
    if not algorithm_rows:
        raise IHCError("Case-marker measurements contain no data rows")
    if any(
        not _is_public_case_alias(str(row.get("case_alias", "")).strip())
        for row in algorithm_rows
    ):
        raise IHCError("Case-marker measurements contain an invalid case_alias")
    algorithm_aliases = {str(row["case_alias"]) for row in algorithm_rows}
    if not set(clinical_by_alias) & algorithm_aliases:
        raise IHCError("No public case aliases overlap between the two inputs")
    algorithm_by_key = {
        (row["case_alias"], row["marker"]): row for row in algorithm_rows
    }
    if len(algorithm_by_key) != len(algorithm_rows):
        raise IHCError("Case-marker measurements contain duplicate case-marker rows")
    comparison_rows: list[dict[str, Any]] = []
    specifications = (
        ("ER", "pathologist_er_percent", "binary-at-1-percent", "none", [0, 1]),
        ("PR", "pathologist_pr_percent", "binary-at-1-percent", "none", [0, 1]),
        (
            "HER2",
            "pathologist_her2_ihc_score",
            "ordinal-0-to-3",
            "quadratic",
            [0, 1, 2, 3],
        ),
        (
            "Ki-67",
            "pathologist_ki67_percent",
            "percentage-deciles",
            "quadratic",
            list(range(10)),
        ),
    )
    summaries: list[dict[str, Any]] = []
    contingency_tables: dict[str, Any] = {}
    for marker, clinical_field, scale, weighting, categories in specifications:
        truth_categories: list[int] = []
        predicted_categories: list[int] = []
        continuous_truth: list[float] = []
        continuous_prediction: list[float] = []
        for alias, clinical in sorted(clinical_by_alias.items()):
            algorithm = algorithm_by_key.get((alias, marker))
            if not algorithm:
                continue
            predicted_value = _number(algorithm.get("marker_pre_score"))
            reference_value = _number(clinical.get(clinical_field))
            if predicted_value is None or reference_value is None:
                continue
            if marker == "HER2":
                if (
                    predicted_value not in categories
                    or reference_value not in categories
                ):
                    raise IHCError(
                        "HER2 agreement values must be exact categories 0, 1, 2, or 3"
                    )
            elif not (0 <= predicted_value <= 100 and 0 <= reference_value <= 100):
                raise IHCError(f"{marker} agreement percentage is outside [0, 100]")
            if marker in {"ER", "PR"}:
                truth_category = int(reference_value >= 1.0)
                predicted_category = int(predicted_value >= 1.0)
            elif marker == "HER2":
                truth_category = int(reference_value)
                predicted_category = int(predicted_value)
            else:
                truth_category = _percentage_decile(reference_value)
                predicted_category = _percentage_decile(predicted_value)
            truth_categories.append(truth_category)
            predicted_categories.append(predicted_category)
            continuous_truth.append(reference_value)
            continuous_prediction.append(predicted_value)
            comparison_rows.append(
                {
                    "case_alias": alias,
                    "marker": marker,
                    "pathologist_value": reference_value,
                    "tumorquantai_value": predicted_value,
                    "pathologist_category": truth_category,
                    "tumorquantai_category": predicted_category,
                    "exact_category_agreement": int(
                        truth_category == predicted_category
                    ),
                }
            )
        kappa, confusion = cohen_kappa(
            truth_categories, predicted_categories, categories, weights=weighting
        )
        low, high = bootstrap_kappa_interval(
            truth_categories,
            predicted_categories,
            categories,
            weights=weighting,
            iterations=bootstrap_iterations,
            seed=bootstrap_seed + IHC_MARKERS.index(marker),
        )
        exact = (
            float(
                np.mean(
                    np.asarray(truth_categories) == np.asarray(predicted_categories)
                )
            )
            if truth_categories
            else math.nan
        )
        mae = (
            float(
                np.mean(
                    np.abs(
                        np.asarray(continuous_truth) - np.asarray(continuous_prediction)
                    )
                )
            )
            if continuous_truth
            else math.nan
        )
        correlation = (
            float(np.corrcoef(continuous_truth, continuous_prediction)[0, 1])
            if len(continuous_truth) >= 2
            and np.std(continuous_truth) > 0
            and np.std(continuous_prediction) > 0
            else math.nan
        )
        summaries.append(
            {
                "marker": marker,
                "primary_scale": scale,
                "weighting": weighting,
                "n": len(truth_categories),
                "kappa": _finite_or_blank(kappa),
                "bootstrap_ci_95_low": _finite_or_blank(low),
                "bootstrap_ci_95_high": _finite_or_blank(high),
                "exact_category_agreement": _finite_or_blank(exact),
                "mean_absolute_error": _finite_or_blank(mae),
                "pearson_correlation": _finite_or_blank(correlation),
            }
        )
        contingency_tables[marker] = {
            "categories": categories,
            "matrix_rows_pathologist_columns_tumorquantai": confusion.tolist(),
        }

    ki_rows = [row for row in comparison_rows if row["marker"] == "Ki-67"]
    ki_truth = [int(float(row["pathologist_value"]) >= 20.0) for row in ki_rows]
    ki_pred = [int(float(row["tumorquantai_value"]) >= 20.0) for row in ki_rows]
    ki_kappa, ki_confusion = cohen_kappa(ki_truth, ki_pred, [0, 1], weights="none")
    ki_low, ki_high = bootstrap_kappa_interval(
        ki_truth,
        ki_pred,
        [0, 1],
        weights="none",
        iterations=bootstrap_iterations,
        seed=bootstrap_seed + 91,
    )
    summaries.append(
        {
            "marker": "Ki-67",
            "primary_scale": "secondary-binary-at-20-percent",
            "weighting": "none",
            "n": len(ki_truth),
            "kappa": _finite_or_blank(ki_kappa),
            "bootstrap_ci_95_low": _finite_or_blank(ki_low),
            "bootstrap_ci_95_high": _finite_or_blank(ki_high),
            "exact_category_agreement": (
                float(np.mean(np.asarray(ki_truth) == np.asarray(ki_pred)))
                if ki_truth
                else ""
            ),
            "mean_absolute_error": "",
            "pearson_correlation": "",
        }
    )
    contingency_tables["Ki-67_binary_20"] = {
        "categories": [0, 1],
        "matrix_rows_pathologist_columns_tumorquantai": ki_confusion.tolist(),
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    summary_fields = (
        "marker",
        "primary_scale",
        "weighting",
        "n",
        "kappa",
        "bootstrap_ci_95_low",
        "bootstrap_ci_95_high",
        "exact_category_agreement",
        "mean_absolute_error",
        "pearson_correlation",
    )
    comparison_fields = (
        "case_alias",
        "marker",
        "pathologist_value",
        "tumorquantai_value",
        "pathologist_category",
        "tumorquantai_category",
        "exact_category_agreement",
    )
    _write_csv(output_dir / "kappa_summary.csv", summaries, summary_fields)
    _write_csv(
        output_dir / "case_comparison_pseudonymized.csv",
        comparison_rows,
        comparison_fields,
    )
    _atomic_json(output_dir / "contingency_tables.json", contingency_tables)
    metadata = {
        "schema_version": "tumorquantai_ihc_agreement_v1",
        "generated_at_utc": utc_now(),
        "bootstrap_iterations": bootstrap_iterations,
        "bootstrap_seed": bootstrap_seed,
        "pathologist_csv_sha256": hashlib.sha256(
            pathologist_csv.read_bytes()
        ).hexdigest(),
        "analysis_case_table_sha256": hashlib.sha256(
            case_path.read_bytes()
        ).hexdigest(),
        "case_rows": len(comparison_rows),
        "paired_marker_rows": len(comparison_rows),
        "pathologist_case_count": len(clinical_by_alias),
        "tumorquantai_case_count": len(algorithm_aliases),
        "overlap_case_count": len(set(clinical_by_alias) & algorithm_aliases),
        "pathologist_only_case_count": len(set(clinical_by_alias) - algorithm_aliases),
        "tumorquantai_only_case_count": len(algorithm_aliases - set(clinical_by_alias)),
        "privacy_status": "pseudonymized_marker_comparison",
        "limitations": [
            "No pathologist-verified invasive-tumour ROI/classifier was supplied.",
            "Selected fields are not whole-slide measurements.",
            "HER2 is a membrane-proxy pre-score, not clinical HER2 status.",
            "Ki-67 20% is a secondary research threshold, not a universal cut point.",
        ],
        "summaries": summaries,
    }
    _atomic_json(output_dir / "agreement_summary.json", metadata)
    table_rows = "".join(
        "<tr>"
        f"<td>{html.escape(str(row['marker']))}</td>"
        f"<td>{html.escape(str(row['primary_scale']))}</td>"
        f"<td>{int(row['n'])}</td>"
        f"<td>{_format_number(row['kappa'], 3)}</td>"
        f"<td>{_format_number(row['bootstrap_ci_95_low'], 3)} to "
        f"{_format_number(row['bootstrap_ci_95_high'], 3)}</td>"
        f"<td>{_format_number(100 * float(row['exact_category_agreement']), 1) + '%' if row['exact_category_agreement'] != '' else '—'}</td></tr>"
        for row in summaries
    )
    report = f"""<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>TumorQuantAI IHC agreement</title><style>body{{max-width:1000px;margin:40px auto;padding:0 22px;font:15px/1.55 system-ui;color:#172033}}h1{{color:#18324a}}.warning{{border-left:5px solid #e4a11b;background:#fff8e8;padding:15px 18px}}table{{width:100%;border-collapse:collapse;margin-top:22px}}th,td{{border-bottom:1px solid #dbe2ea;padding:10px;text-align:left}}th{{background:#f3f6f9}}code{{font-size:12px}}</style></head><body><h1>Pathologist–TumorQuantAI concordance</h1><div class="warning"><b>Research use only.</b> Kappa measures agreement under the prespecified categories; it does not validate the image method or establish clinical accuracy. Review sampling, segmentation overlays, category prevalence, and confidence intervals.</div><table><thead><tr><th>Marker</th><th>Scale</th><th>n</th><th>κ</th><th>Bootstrap 95% CI</th><th>Exact agreement</th></tr></thead><tbody>{table_rows}</tbody></table><p><a href="kappa_summary.csv">Kappa CSV</a> · <a href="contingency_tables.json">Contingency tables</a> · <a href="case_comparison_pseudonymized.csv">Pseudonymized paired values</a></p><p>Method reference: <a href="https://doi.org/10.1177/001316446002000104">Cohen (1960), A coefficient of agreement for nominal scales</a>. Clinical interpretation context: <a href="https://ascopubs.org/doi/10.1200/JCO.19.02309">ER/PR</a>, <a href="https://www.cap.org/cap-guidelines/her2-testing-in-breast-cancer-2023-guideline-update/">HER2</a>, and <a href="https://pmc.ncbi.nlm.nih.gov/articles/PMC8487652/">Ki-67</a> guidance.</p></body></html>"""
    _atomic_text(output_dir / "AGREEMENT_REPORT.html", report)
    return {**metadata, "report_path": str(output_dir / "AGREEMENT_REPORT.html")}


def add_cli_parser(subparsers: Any) -> None:
    parser = subparsers.add_parser(
        "ihc", help="quantify breast IHC markers and assess pathologist agreement"
    )
    commands = parser.add_subparsers(dest="ihc_command", required=True)
    quantify = commands.add_parser(
        "quantify",
        help="segment and quantify ER, PR, HER2, and Ki-67 TIFF patches",
    )
    quantify.add_argument(
        "input",
        type=Path,
        help="extracted patch root or case-archive directory",
    )
    quantify.add_argument(
        "--manifest", required=True, type=Path, help="public patch_manifest.csv"
    )
    quantify.add_argument(
        "--output",
        required=True,
        type=Path,
        help="new or resumable IHC result directory",
    )
    quantify.add_argument(
        "--markers", nargs="+", choices=IHC_MARKERS, default=list(IHC_MARKERS)
    )
    quantify.add_argument("--include", default="*", help="case-alias glob to include")
    quantify.add_argument("--exclude", default="", help="case-alias glob to exclude")
    quantify.add_argument("--workers", type=int, default=min(12, os.cpu_count() or 1))
    quantify.add_argument(
        "--allow-missing",
        action="store_true",
        help="record unavailable selected patches instead of failing closed",
    )
    quantify.add_argument("--fail-fast", action="store_true")
    quantify.add_argument(
        "--save-cells",
        action="store_true",
        help="write per-cell gzip CSV tables",
    )
    quantify.add_argument(
        "--no-qc", action="store_true", help="skip PNG segmentation overlays"
    )
    quantify.add_argument(
        "--no-verify-decoded-rgb",
        action="store_true",
        help="skip manifest decoded-pixel hashes",
    )
    quantify.add_argument(
        "--no-resume",
        action="store_true",
        help="recompute patches instead of reusing matching records",
    )
    quantify.add_argument("--dry-run", action="store_true")
    quantify.add_argument("--weak-dab-od", type=float, default=0.20)
    quantify.add_argument("--moderate-dab-od", type=float, default=0.40)
    quantify.add_argument("--strong-dab-od", type=float, default=0.60)
    quantify.add_argument("--minimum-cells-for-score", type=int, default=100)

    clinical = commands.add_parser(
        "anonymize-clinical",
        help="export the minimum pathologist marker table under public aliases",
    )
    clinical.add_argument("workbook", type=Path)
    clinical.add_argument(
        "--linkage",
        required=True,
        type=Path,
        help="private case_alias-to-clinical-ID linkage; never published",
    )
    clinical.add_argument("--output", required=True, type=Path)
    clinical.add_argument("--sheet", default="Biopsias finales incluidas")
    clinical.add_argument("--clinical-id-column")
    clinical.add_argument("--linkage-id-column")

    compare = commands.add_parser(
        "compare",
        help="calculate marker-wise kappa against a privacy-minimized CSV",
    )
    compare.add_argument("results", type=Path, help="TumorQuantAI IHC result directory")
    compare.add_argument("--pathologist-csv", required=True, type=Path)
    compare.add_argument("--output", type=Path, help="default: RESULTS/agreement")
    compare.add_argument("--bootstrap-iterations", type=int, default=2000)
    compare.add_argument("--bootstrap-seed", type=int, default=20260829)


def dispatch_cli(args: Any) -> int:
    if np is None:
        raise IHCError(
            "NumPy is required for IHC commands; run tumorquantai install first"
        )
    if args.ihc_command == "quantify":
        config = IHCConfig(
            weak_dab_od=args.weak_dab_od,
            moderate_dab_od=args.moderate_dab_od,
            strong_dab_od=args.strong_dab_od,
            minimum_cells_for_score=args.minimum_cells_for_score,
        )
        records, unavailable = load_patch_manifest(
            args.manifest,
            args.input,
            markers=args.markers,
            include=args.include,
            exclude=args.exclude,
        )
        result = run_quantification(
            records,
            unavailable,
            args.manifest,
            args.input,
            args.output,
            config,
            workers=args.workers,
            verify_decoded_rgb=not args.no_verify_decoded_rgb,
            save_cells=args.save_cells,
            save_qc=not args.no_qc,
            resume=not args.no_resume,
            allow_missing=args.allow_missing,
            fail_fast=args.fail_fast,
            dry_run=args.dry_run,
        )
        print(json.dumps(result, indent=2, sort_keys=True))
        return 0 if result.get("failed_patch_count", 0) == 0 else 10
    if args.ihc_command == "anonymize-clinical":
        result = export_pathologist_csv(
            args.workbook,
            args.linkage,
            args.output,
            sheet_name=args.sheet,
            clinical_id_column=args.clinical_id_column,
            linkage_id_column=args.linkage_id_column,
        )
        print(json.dumps(result, indent=2, sort_keys=True))
        return 0
    if args.ihc_command == "compare":
        results = args.results.expanduser().resolve()
        output = (
            args.output.expanduser().resolve() if args.output else results / "agreement"
        )
        result = compare_pathologist_agreement(
            results,
            args.pathologist_csv.expanduser().resolve(),
            output,
            bootstrap_iterations=args.bootstrap_iterations,
            bootstrap_seed=args.bootstrap_seed,
        )
        print(json.dumps(result, indent=2, sort_keys=True))
        return 0
    raise IHCError("Unknown IHC command")
